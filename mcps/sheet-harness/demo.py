#!/usr/bin/env python3
"""
Demo script showcasing TAIA Sheet Harness capabilities.
Run with: .venv/Scripts/python demo.py
"""

import io
import base64
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# Create a sample test workbook
def create_sample_workbook():
    """Create a realistic test workbook with multiple test cases."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create required tabs
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    
    # Main test sheet
    main = wb.create_sheet("TPI-TP-443 - User Dashboard Tests")
    
    # Header row
    headers = ["No", "Prerequisite", "Test Detail", "Expected Results", 
               "Data No", "Remarks", "Num Cases", "Test Date", "Test OK/NG", "Tester"]
    for col, h in enumerate(headers, 1):
        main.cell(row=1, column=col, value=h)
    
    # Test case 1: Login
    main.cell(row=2, column=1, value=1)
    main.cell(row=2, column=2, value="User has valid credentials")
    main.cell(row=2, column=3, value="Navigate to login page and enter credentials")
    main.cell(row=2, column=4, value="User is redirected to dashboard")
    main.cell(row=2, column=5, value="TC_LOGIN_001")
    main.cell(row=2, column=7, value="1")
    
    # Test case 2: Dashboard load
    main.cell(row=3, column=1, value=2)
    main.cell(row=3, column=2, value="User is logged in")
    main.cell(row=3, column=3, value="Navigate to dashboard page")
    main.cell(row=3, column=4, value="Dashboard loads with all widgets")
    main.cell(row=3, column=5, value="TC_DASH_001")
    main.cell(row=3, column=7, value="1")
    
    # Test case 3: Settings panel
    main.cell(row=4, column=1, value=3)
    main.cell(row=4, column=2, value="On dashboard page")
    main.cell(row=4, column=3, value="Click settings gear icon")
    main.cell(row=4, column=4, value="Settings panel slides out")
    main.cell(row=4, column=5, value="TC_SETTINGS_001")
    main.cell(row=4, column=7, value="1")
    
    # Test case 4: Logout
    main.cell(row=5, column=1, value=4)
    main.cell(row=5, column=2, value="User is on dashboard")
    main.cell(row=5, column=3, value="Click logout button")
    main.cell(row=5, column=4, value="User is redirected to login page")
    main.cell(row=5, column=5, value="TC_LOGOUT_001")
    main.cell(row=5, column=7, value="1")
    
    # Create screenshot tabs
    for test_no in [2, 3, 4]:
        tab = wb.create_sheet(f"No.{test_no}")
        tab.cell(row=1, column=2, value="Test No.")
        tab.cell(row=2, column=2, value=test_no)
    
    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def create_test_screenshot():
    """Create a simple test screenshot."""
    img = PILImage.new('RGB', (800, 600), color=(73, 109, 137))
    bio = io.BytesIO()
    img.save(bio, format='PNG')
    return base64.b64encode(bio.getvalue()).decode('ascii')

def demo():
    print("=" * 70)
    print("TAIA SHEET HARNESS - DEMONSTRATION")
    print("=" * 70)
    
    # Import the harness modules
    from taia_sheet_harness.server import (
        import_workbook_tool,
        get_test_cases,
        write_result_tool,
        embed_screenshot_tool,
        get_workbook_summary_tool,
    )
    from taia_sheet_harness.models import OkNg
    
    print("\n[Step 1] Creating sample test workbook...")
    wb_bytes = create_sample_workbook()
    wb_b64 = base64.b64encode(wb_bytes).decode('ascii')
    print(f"  Workbook size: {len(wb_bytes):,} bytes")
    print(f"  Contains: 4 test cases, 3 screenshot tabs")
    
    print("\n[Step 2] Importing workbook...")
    result = import_workbook_tool(
        file_b64=wb_b64,
        filename="dashboard_tests.xlsx",
        overwrite=True
    )
    print(f"  [OK] Imported: {result['filename']}")
    print(f"  [OK] Test cases: {result['test_count']}")
    print(f"  [OK] OK: {result['ok_count']}, NG: {result['ng_count']}, Untested: {result['untested_count']}")
    
    print("\n[Step 3] Reading test cases...")
    result = get_test_cases(workbook_id="dashboard_tests")
    print(f"  [OK] Retrieved {result['count']} test cases:")
    for tc in result['test_cases']:
        status = f"{tc['test_okng']}" if tc['test_okng'] else "UNTESTED"
        print(f"    Row {tc['row']}: Test #{tc['test_no']} - {tc['test_detail'][:40]}... [{status}]")
    
    print("\n[Step 4] Running tests and writing results...")
    
    # Test 1: PASS
    result = write_result_tool(
        workbook_id="dashboard_tests",
        row=2,
        ok_ng=OkNg.OK,
        notes="Login successful, redirected in 1.2s"
    )
    print(f"  [OK] Test #1 (Login): {result['ok_ng']}")
    
    # Test 2: PASS
    result = write_result_tool(
        workbook_id="dashboard_tests",
        row=3,
        ok_ng=OkNg.OK,
        notes="All widgets loaded correctly"
    )
    print(f"  [OK] Test #2 (Dashboard): {result['ok_ng']}")
    
    # Test 3: FAIL
    result = write_result_tool(
        workbook_id="dashboard_tests",
        row=4,
        ok_ng=OkNg.NG,
        notes="Settings panel timeout after 5s"
    )
    print(f"  [OK] Test #3 (Settings): {result['ok_ng']}")
    
    # Test 4: PASS
    result = write_result_tool(
        workbook_id="dashboard_tests",
        row=5,
        ok_ng=OkNg.OK,
        notes="Logout successful, session cleared"
    )
    print(f"  [OK] Test #4 (Logout): {result['ok_ng']}")
    
    print("\n[Step 5] Embedding screenshots for failures...")
    screenshot_b64 = create_test_screenshot()
    
    result = embed_screenshot_tool(
        workbook_id="dashboard_tests",
        test_no=3,
        image_b64=screenshot_b64,
        caption="Settings panel did not open - timeout error",
        step_number=1
    )
    print(f"  [OK] Screenshot embedded in tab: {result['tab_name']}")
    
    print("\n[Step 6] Getting updated summary...")
    result = get_workbook_summary_tool(workbook_id="dashboard_tests")
    print(f"  [OK] Total tests: {result['test_count']}")
    print(f"  [OK] Passed: {result['ok_count']}")
    print(f"  [OK] Failed: {result['ng_count']}")
    print(f"  [OK] Untested: {result['untested_count']}")
    print(f"  [OK] Screenshot tabs: {result['screenshot_tabs']}")
    
    print("\n[Step 7] Access updated workbook via TAIA chat")
    print(f"  [OK] Download: Available through TAIA file browser")
    
    print("\n[Step 8] Reading final test cases...")
    result = get_test_cases(workbook_id="dashboard_tests")
    print(f"  [OK] Final status:")
    for tc in result['test_cases']:
        status = f"{tc['test_okng']}" if tc['test_okng'] else "UNTESTED"
        notes = f" - {tc['remarks']}" if tc['remarks'] else ""
        screenshot = f" [screenshot: {tc['screenshot_tab']}]" if tc['screenshot_tab'] else ""
        print(f"    Test #{tc['test_no']}: {status}{notes}{screenshot}")
    
    print("\n" + "=" * 70)
    print("DEMONSTRATION COMPLETE")
    print("=" * 70)
    print("\nThe workbook has been updated with:")
    print("  • Test results (OK/NG) with color coding")
    print("  • Test dates and tester name")
    print("  • Failure notes in Remarks column")
    print("  • Embedded screenshots in screenshot tabs")
    print("\nDownload the updated file through TAIA's MCP integration")
    print("  or access directly from the workbooks volume")
    print("=" * 70)

if __name__ == "__main__":
    demo()