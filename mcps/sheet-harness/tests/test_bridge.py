import pytest
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from taia_sheet_harness.taia_bridge import (
    taia_import_workbook, taia_get_test_cases, taia_write_result,
    taia_embed_screenshot, taia_get_workbook_meta
)
from taia_sheet_harness.models import OkNg


@pytest.fixture
def tmp_bridge_workbooks_dir(tmp_path):
    """Override WORKBOOKS_DIR for bridge testing."""
    import taia_sheet_harness.server
    import taia_sheet_harness.taia_bridge
    original = taia_sheet_harness.server.WORKBOOKS_DIR
    taia_sheet_harness.server.WORKBOOKS_DIR = tmp_path
    taia_sheet_harness.taia_bridge.WORKBOOKS_DIR = tmp_path
    yield tmp_path
    taia_sheet_harness.server.WORKBOOKS_DIR = original
    taia_sheet_harness.taia_bridge.WORKBOOKS_DIR = original


def test_taia_import_workbook(tmp_bridge_workbooks_dir):
    """Test bridge import function."""
    from openpyxl import Workbook
    import io

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value=1)
    wb.create_sheet("No.2")

    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    result = taia_import_workbook("test.xlsx", file_bytes)
    assert result["success"] is True
    assert result["workbook_id"] == "test"
    assert result["test_count"] == 1
    assert "No.2" in result["screenshot_tabs"]


def test_taia_get_test_cases(tmp_bridge_workbooks_dir):
    """Test bridge get test cases function."""
    from openpyxl import Workbook
    import io

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="Test")
    wb.create_sheet("No.2")

    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    # Import first
    taia_import_workbook("test.xlsx", file_bytes)

    # Get test cases
    result = taia_get_test_cases("test")
    assert result["success"] is True
    assert result["count"] == 1
    assert len(result["test_cases"]) == 1
    assert result["test_cases"][0]["test_no"] == 1


def test_taia_write_result(tmp_bridge_workbooks_dir):
    """Test bridge write result function."""
    from openpyxl import Workbook
    import io

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    wb.create_sheet("No.2")

    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    # Import first
    taia_import_workbook("test.xlsx", file_bytes)

    # Write result
    result = taia_write_result("test", 2, "OK", "Test passed")
    assert result["success"] is True
    assert result["ok_ng"] == "OK"

    # Verify by reading
    cases = taia_get_test_cases("test")
    assert cases["test_cases"][0]["test_okng"] == "OK"


def test_taia_embed_screenshot(tmp_bridge_workbooks_dir):
    """Test bridge embed screenshot function."""
    from openpyxl import Workbook
    import io
    import base64

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")

    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    # Import first
    taia_import_workbook("test.xlsx", file_bytes)

    # Create a simple test image (1x1 pixel PNG)
    png_1x1 = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="

    # Embed screenshot
    result = taia_embed_screenshot("test", 2, png_1x1, "Test screenshot", 1)
    assert result["success"] is True
    assert result["tab_name"] == "No.2"


def test_taia_get_workbook_meta(tmp_bridge_workbooks_dir):
    """Test bridge get workbook meta function."""
    from openpyxl import Workbook
    import io

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value=1)
    wb.create_sheet("No.2")

    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    # Import first
    taia_import_workbook("test.xlsx", file_bytes)

    # Get meta
    result = taia_get_workbook_meta("test")
    assert result["success"] is True
    assert result["test_count"] == 1
    assert result["untested_count"] == 1
    assert "No.2" in result["screenshot_tabs"]


def test_taia_write_result_validation(tmp_bridge_workbooks_dir):
    """Test bridge write result validation."""
    result = taia_write_result("nonexistent", 2, "OK")
    assert result["success"] is False

    result = taia_write_result("test", 1, "OK")  # Invalid row
    assert result["success"] is False
    assert "Row must be >= 2" in result["error"]

    result = taia_write_result("test", 2, "INVALID")  # Invalid OK/NG
    assert result["success"] is False
    assert "must be 'OK' or 'NG'" in result["error"]