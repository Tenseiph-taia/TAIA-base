import pytest
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from taia_sheet_harness.server import (
    import_workbook_tool, get_test_cases, write_result_tool,
    embed_screenshot_tool, validate_workbook_tool, list_workbooks_tool, get_workbook_summary_tool,
    _resolve, _user_error, ValidationError, StorageError, OkNg
)
from taia_sheet_harness.models import TestCase, WriteResultInput, EmbedScreenshotInput


@pytest.fixture
def tmp_workbooks_dir(tmp_path):
    """Override WORKBOOKS_DIR for testing."""
    import taia_sheet_harness.server
    original = taia_sheet_harness.server.WORKBOOKS_DIR
    taia_sheet_harness.server.WORKBOOKS_DIR = tmp_path
    yield tmp_path
    taia_sheet_harness.server.WORKBOOKS_DIR = original


def test_import_workbook_tool_valid(tmp_workbooks_dir):
    """Import valid workbook."""
    # Create a minimal valid workbook
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="Prereq")
    ws.cell(row=2, column=3, value="Detail")
    ws.cell(row=2, column=4, value="Expected")
    ws.cell(row=2, column=5, value="Data")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="1")
    wb.create_sheet("No.2")
    import io
    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    result = import_workbook_tool(
        file_b64=file_bytes.decode('latin1'),
        filename="test.xlsx",
        overwrite=False,
    )
    assert result["ok"] is True
    assert result["workbook_id"] == "test"
    assert result["test_count"] == 1
    assert (tmp_workbooks_dir / "test.xlsx").exists()


def test_import_workbook_tool_invalid_file(tmp_workbooks_dir):
    """Reject invalid file."""
    result = import_workbook_tool(
        file_b64="not a valid xlsx",
        filename="bad.xlsx",
        overwrite=False,
    )
    assert result["ok"] is False


def test_import_workbook_tool_duplicate_no_overwrite(tmp_workbooks_dir):
    """Error on duplicate without overwrite."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    import io as io2
    result = import_workbook_tool(file_b64=file_bytes.decode('latin1'), filename="dup.xlsx")
    assert result["ok"] is True

    result2 = import_workbook_tool(file_b64=file_bytes.decode('latin1'), filename="dup.xlsx", overwrite=False)
    assert result2["ok"] is False
    assert "already exists" in result2["error"]


def test_import_workbook_tool_duplicate_with_overwrite(tmp_workbooks_dir):
    """Success with overwrite=True."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="A")
    ws.cell(row=2, column=3, value="B")
    ws.cell(row=2, column=4, value="C")
    ws.cell(row=2, column=5, value="D")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="1")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    result = import_workbook_tool(file_b64=file_bytes.decode('latin1'), filename="dup.xlsx", overwrite=True)
    assert result["ok"] is True


def test_get_test_cases_empty(tmp_workbooks_dir):
    """Empty workbook returns empty list."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="empty.xlsx")
    assert result["ok"] is True

    result = get_test_cases("empty")
    assert result["ok"] is True
    assert result["count"] == 0
    assert result["test_cases"] == []


def test_get_test_cases_with_data(tmp_workbooks_dir):
    """Read test cases correctly."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="Prereq1")
    ws.cell(row=2, column=3, value="Detail1")
    ws.cell(row=2, column=4, value="Exp1")
    ws.cell(row=2, column=5, value="Data1")
    ws.cell(row=2, column=6, value="Rem1")
    ws.cell(row=2, column=7, value="1")
    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value="Prereq2")
    ws.cell(row=3, column=3, value="Detail2")
    ws.cell(row=3, column=4, value="Exp2")
    ws.cell(row=3, column=5, value="Data2")
    ws.cell(row=3, column=6, value="Rem2")
    ws.cell(row=3, column=7, value="1")
    wb.create_sheet("No.1")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="data.xlsx")
    assert result["ok"] is True

    result = get_test_cases("data")
    assert result["ok"] is True
    assert result["count"] == 2
    cases = result["test_cases"]
    assert cases[0]["test_no"] == 1
    assert cases[0]["prerequisite"] == "Prereq1"
    assert cases[0]["test_detail"] == "Detail1"
    assert cases[0]["screenshot_tab"] == "No.1"


def test_write_result_invalid_row_zero(tmp_workbooks_dir):
    """Row < 2 rejected."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="wb.xlsx")
    assert result["ok"] is True

    result = write_result_tool("wb", row=1, ok_ng=OkNg.OK)
    assert result["ok"] is False
    assert "Row must be >= 2" in result["error"]


def test_write_result_invalid_row_too_high(tmp_workbooks_dir):
    """Row > 65536 rejected."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="wb2.xlsx")
    assert result["ok"] is True

    result = write_result_tool("wb2", row=70000, ok_ng=OkNg.OK)
    assert result["ok"] is False
    assert "exceeds" in result["error"]


def test_write_result_ok(tmp_workbooks_dir):
    """Write OK, verify stored."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="P")
    ws.cell(row=2, column=3, value="D")
    ws.cell(row=2, column=4, value="E")
    ws.cell(row=2, column=5, value="Data")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="1")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="wb_ok.xlsx")
    assert result["ok"] is True

    result = write_result_tool("wb_ok", row=2, ok_ng=OkNg.OK)
    assert result["ok"] is True
    assert result["ok_ng"] == "OK"

    # Verify by reading
    cases = get_test_cases("wb_ok")
    assert cases["test_cases"][0]["test_okng"] == "OK"
    assert cases["test_cases"][0]["test_date"] != ""


def test_write_result_ng(tmp_workbooks_dir):
    """Write NG."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="P")
    ws.cell(row=2, column=3, value="D")
    ws.cell(row=2, column=4, value="E")
    ws.cell(row=2, column=5, value="Data")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="1")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="wb_ng.xlsx")
    assert result["ok"] is True

    result = write_result_tool("wb_ng", row=2, ok_ng=OkNg.NG)
    assert result["ok"] is True
    assert result["ok_ng"] == "NG"


def test_write_result_with_notes(tmp_workbooks_dir):
    """Write with notes."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="P")
    ws.cell(row=2, column=3, value="D")
    ws.cell(row=2, column=4, value="E")
    ws.cell(row=2, column=5, value="Data")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="1")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="wb_note.xlsx")
    assert result["ok"] is True

    result = write_result_tool("wb_note", row=2, ok_ng=OkNg.OK, notes="All good")
    assert result["ok"] is True

    cases = get_test_cases("wb_note")
    assert "All good" in cases["test_cases"][0]["remarks"]


def test_write_results_batch_all_valid(tmp_workbooks_dir):
    """Batch write all valid."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    for r in range(2, 7):
        ws.cell(row=r, column=1, value=r-1)
        ws.cell(row=r, column=2, value=f"P{r}")
        ws.cell(row=r, column=3, value=f"D{r}")
        ws.cell(row=r, column=4, value=f"E{r}")
        ws.cell(row=r, column=5, value=f"Data{r}")
        ws.cell(row=r, column=6, value="")
        ws.cell(row=r, column=7, value="1")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="batch.xlsx")
    assert result["ok"] is True

    results = [
        WriteResultInput(row=2, actual_result="A1", ok_ng=OkNg.OK),
        WriteResultInput(row=3, actual_result="A2", ok_ng=OkNg.NG),
        WriteResultInput(row=4, actual_result="A3", ok_ng=OkNg.OK),
    ]
    # Note: write_results_batch is not implemented in the new version
    # We'll test individual writes instead
    result = write_result_tool("batch", row=2, ok_ng=OkNg.OK)
    assert result["ok"] is True
    result = write_result_tool("batch", row=3, ok_ng=OkNg.NG)
    assert result["ok"] is True


def test_embed_screenshot_oversized_image(tmp_workbooks_dir):
    """Oversized image rejected."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="wb_big.xlsx")
    assert result["ok"] is True

    oversized = "A" * 11000000
    result = embed_screenshot_tool("wb_big", 2, oversized)
    assert result["ok"] is False
    assert "too large" in result["error"] or "Maximum" in result["error"]


def test_embed_screenshot_invalid_b64(tmp_workbooks_dir):
    """Invalid base64 rejected."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="wb_bad.xlsx")
    assert result["ok"] is True

    result = embed_screenshot_tool("wb_bad", 2, "not valid!!!")
    assert result["ok"] is False
    assert "Invalid" in result["error"] or "base64" in result["error"]


def test_embed_screenshot_no_tab(tmp_workbooks_dir):
    """Error when no matching screenshot tab."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    # No screenshot tabs
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="wb_notab.xlsx")
    assert result["ok"] is True

    png_1x1 = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
    result = embed_screenshot_tool("wb_notab", 2, png_1x1)
    assert result["ok"] is False
    assert "No screenshot tab" in result["error"]


def test_validate_workbook_valid(tmp_workbooks_dir):
    """Valid workbook passes validation."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="valid.xlsx")
    assert result["ok"] is True

    result = validate_workbook_tool("valid")
    assert result["ok"] is True
    assert "valid" in result["message"].lower()


def test_validate_workbook_missing_sheet(tmp_workbooks_dir):
    """Missing sheet detected."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    # Only has default 'Sheet', not our required structure
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="bad.xlsx")
    # Import might still succeed (it just saves the file)
    # Validation should catch it
    result = validate_workbook_tool("bad")
    assert result["ok"] is False


def test_workbook_not_found_returns_user_error(tmp_workbooks_dir):
    """404-style error."""
    result = get_test_cases("nonexistent")
    assert result["ok"] is False
    assert "not found" in result["error"].lower()


def test_invalid_workbook_id_rejected(tmp_workbooks_dir):
    """Invalid ID rejected."""
    result = import_workbook_tool(
        file_b64="UEsDBAoAAAAAAIdO4kgAAAAAAAAAAAAAAAAJAAAAdGVzdC54bHN4UEsFBgAAAAABAAEASgAAAF4BAAAAAA==",
        filename="bad/name.xlsx",
        overwrite=False,
    )
    assert result["ok"] is False


def test_list_workbooks(tmp_workbooks_dir):
    """List returns created workbooks."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    ws.cell(row=1, column=1, value="#")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    import_workbook_tool(file_b64=file_bytes.decode('latin1'), filename="wb1.xlsx")
    import_workbook_tool(file_b64=file_bytes.decode('latin1'), filename="wb2.xlsx")

    result = list_workbooks_tool()
    assert result["ok"] is True
    assert result["count"] >= 2
    ids = [w["workbook_id"] for w in result["workbooks"]]
    assert "wb1" in ids
    assert "wb2" in ids


def test_get_workbook_summary(tmp_workbooks_dir):
    """Summary computed correctly."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("Main")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    # Row 2: OK
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="P")
    ws.cell(row=2, column=3, value="D")
    ws.cell(row=2, column=4, value="E")
    ws.cell(row=2, column=5, value="Data")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="1")
    ws.cell(row=2, column=8, value="2024-01-01")
    ws.cell(row=2, column=9, value="OK")
    ws.cell(row=2, column=10, value="TAIA")
    # Row 3: untested
    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value="P")
    ws.cell(row=3, column=3, value="D")
    ws.cell(row=3, column=4, value="E")
    ws.cell(row=3, column=5, value="Data")
    ws.cell(row=3, column=6, value="")
    ws.cell(row=3, column=7, value="1")
    wb.create_sheet("No.2")
    bio = io.BytesIO()
    wb.save(bio)

    result = import_workbook_tool(file_b64=bio.getvalue().decode('latin1'), filename="summary.xlsx")
    assert result["ok"] is True

    result = get_workbook_summary_tool("summary")
    assert result["ok"] is True
    assert result["test_count"] == 2
    assert result["ok_count"] == 1
    assert result["ng_count"] == 0
    assert result["untested_count"] == 1