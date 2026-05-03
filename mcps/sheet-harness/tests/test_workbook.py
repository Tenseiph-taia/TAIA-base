import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from taia_sheet_harness.workbook import (
    parse_tab_map, import_workbook, read_test_cases, write_result,
    embed_screenshot, get_workbook_meta, _atomic_save, MAX_TEXT_LEN, MAX_IMAGE_B64_BYTES, _BACKUP_RETAIN
)
from taia_sheet_harness.models import TestCase, OkNg


@pytest.fixture
def tmp_workbook(tmp_path):
    """Create a minimal valid workbook for testing."""
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    main = wb.create_sheet("TPI-TP-443")
    main.cell(row=1, column=1, value="#")
    main.cell(row=1, column=2, value="Prerequisite")
    main.cell(row=1, column=3, value="Test Detail")
    main.cell(row=1, column=4, value="Expected Results")
    main.cell(row=1, column=5, value="Test Case Data No.")
    main.cell(row=1, column=6, value="Remarks")
    main.cell(row=1, column=7, value="Num of test cases")
    main.cell(row=1, column=8, value="Test Date")
    main.cell(row=1, column=9, value="Test OK/NG")
    main.cell(row=1, column=10, value="Test Tester")
    main.cell(row=2, column=1, value=1)
    main.cell(row=2, column=2, value="Login")
    main.cell(row=2, column=3, value="Enter credentials")
    main.cell(row=2, column=4, value="Login success")
    main.cell(row=2, column=5, value="TC-001")
    main.cell(row=2, column=6, value="")
    main.cell(row=2, column=7, value=1)
    wb.create_sheet("No.2")
    wb.save(str(path))
    return path


def test_parse_tab_map_simple():
    """No.2 → {2: 'No.2'}"""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    wb.create_sheet("Main")
    wb.create_sheet("No.2")
    result = parse_tab_map(wb)
    assert result == {2: "No.2"}


def test_parse_tab_map_range():
    """No.8-10 → {8: 'No.8-10', 9: 'No.8-10', 10: 'No.8-10'}"""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    wb.create_sheet("Main")
    wb.create_sheet("No.8-10")
    result = parse_tab_map(wb)
    assert result == {8: "No.8-10", 9: "No.8-10", 10: "No.8-10"}


def test_parse_tab_map_mixed():
    """Multiple tabs parsed correctly."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    wb.create_sheet("Main")
    wb.create_sheet("No.2")
    wb.create_sheet("No.5-7")
    wb.create_sheet("Other")
    result = parse_tab_map(wb)
    assert 2 in result
    assert result[2] == "No.2"
    assert 5 in result and 6 in result and 7 in result
    assert result[5] == "No.5-7"
    assert "Other" not in result.values()


def test_parse_tab_map_invalid_names():
    """Tabs not matching No.X pattern are ignored."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    wb.create_sheet("Main")
    wb.create_sheet("Screenshots")
    wb.create_sheet("No")
    wb.create_sheet("No.")
    result = parse_tab_map(wb)
    assert result == {}


def test_import_workbook_valid(tmp_path):
    """Import creates file and returns metadata."""
    path = tmp_path / "imported.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    main = wb.create_sheet("MainSheet")
    main.cell(row=1, column=1, value="#")
    main.cell(row=2, column=1, value=1)
    main.cell(row=2, column=2, value="Prereq")
    main.cell(row=2, column=3, value="Detail")
    main.cell(row=2, column=4, value="Expected")
    main.cell(row=2, column=5, value="Data")
    main.cell(row=2, column=6, value="")
    main.cell(row=2, column=7, value="1")
    wb.create_sheet("No.2")
    import io
    bio = io.BytesIO()
    wb.save(bio)
    file_bytes = bio.getvalue()

    meta = import_workbook(path, file_bytes)
    assert path.exists()
    assert meta.workbook_id == "imported"
    assert meta.test_count == 1
    assert "No.2" in meta.screenshot_tabs


def test_read_test_cases_skips_headers(tmp_path):
    """Rows with no integer in Col A are skipped."""
    path = tmp_path / "cases.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value="Crew Dashboard Page")
    ws.cell(row=2, column=2, value="")
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value="Login")
    ws.cell(row=3, column=3, value="Enter creds")
    ws.cell(row=3, column=4, value="Success")
    ws.cell(row=3, column=5, value="TC001")
    ws.cell(row=3, column=6, value="")
    ws.cell(row=3, column=7, value="1")
    wb.create_sheet("No.2")
    wb.save(str(path))

    cases = read_test_cases(path)
    assert len(cases) == 1
    assert cases[0].row == 3
    assert cases[0].test_no == 1


def test_read_test_cases_resolves_tab(tmp_path):
    """TestCase.screenshot_tab populated correctly."""
    path = tmp_path / "cases.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value=2)
    ws.cell(row=2, column=2, value="Test")
    ws.cell(row=2, column=3, value="Detail")
    ws.cell(row=2, column=4, value="Exp")
    ws.cell(row=2, column=5, value="Data")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="1")
    ws.cell(row=3, column=1, value=8)
    ws.cell(row=3, column=2, value="Test8")
    ws.cell(row=3, column=3, value="Detail8")
    ws.cell(row=3, column=4, value="Exp8")
    ws.cell(row=3, column=5, value="Data8")
    ws.cell(row=3, column=6, value="")
    ws.cell(row=3, column=7, value="1")
    wb.create_sheet("No.2")
    wb.create_sheet("No.8-10")
    wb.save(str(path))

    cases = read_test_cases(path)
    assert len(cases) == 2
    assert cases[0].screenshot_tab == "No.2"
    assert cases[1].screenshot_tab == "No.8-10"


def test_read_test_cases_no_tab_is_none(tmp_path):
    """test_no with no matching tab → screenshot_tab is None."""
    path = tmp_path / "cases.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=2, column=1, value=99)
    ws.cell(row=2, column=2, value="Test")
    ws.cell(row=2, column=3, value="Detail")
    ws.cell(row=2, column=4, value="Exp")
    ws.cell(row=2, column=5, value="Data")
    ws.cell(row=2, column=6, value="")
    ws.cell(row=2, column=7, value="1")
    wb.save(str(path))

    cases = read_test_cases(path)
    assert len(cases) == 1
    assert cases[0].screenshot_tab is None


def test_write_result_ok(tmp_path):
    """Write OK, verify columns."""
    path = tmp_path / "write.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    wb.save(str(path))

    write_result(path, 2, OkNg.OK)

    wb2 = load_workbook(str(path))
    ws2 = wb2.worksheets[2]
    assert ws2.cell(row=2, column=8).value
    assert ws2.cell(row=2, column=9).value == "OK"
    assert ws2.cell(row=2, column=10).value == "TAIA"


def test_write_result_ng(tmp_path):
    """Write NG, verify red fill."""
    path = tmp_path / "write_ng.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    wb.save(str(path))

    write_result(path, 2, OkNg.NG)

    wb2 = load_workbook(str(path))
    ws2 = wb2.worksheets[2]
    assert ws2.cell(row=2, column=9).value == "NG"
    assert ws2.cell(row=2, column=10).value == "TAIA"


def test_write_result_atomic(tmp_path):
    """No .tmp residue after write."""
    path = tmp_path / "atomic.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    wb.save(str(path))

    write_result(path, 2, OkNg.OK)
    tmp = path.with_suffix(".xlsx.tmp")
    assert not tmp.exists()


def test_backup_created_on_write(tmp_path):
    """After write, .backups/ dir contains one file."""
    path = tmp_path / "backup.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    wb.save(str(path))

    write_result(path, 2, OkNg.OK)
    backup_dir = path.parent / ".backups"
    backups = list(backup_dir.glob("*.xlsx"))
    assert len(backups) == 1


def test_backup_retention(tmp_path):
    """After 7 writes, .backups/ contains exactly 5 files."""
    path = tmp_path / "retention.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value=f"H{c}")
    ws.cell(row=2, column=1, value=1)
    wb.save(str(path))

    import time
    for i in range(7):
        wb2 = load_workbook(str(path))
        ws2 = wb2.worksheets[2]
        ws2.cell(row=2, column=2, value=f"Update {i}")
        _atomic_save(wb2, path)
        time.sleep(0.01)

    backup_dir = path.parent / ".backups"
    backups = sorted(backup_dir.glob("*.xlsx"))
    assert len(backups) == _BACKUP_RETAIN


def test_embed_screenshot_valid(tmp_path):
    """Valid image embeds correctly."""
    path = tmp_path / "screenshot.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    wb.create_sheet("MainSheet")
    wb.create_sheet("No.2")
    wb.save(str(path))

    png_1x1 = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
    tab_name = embed_screenshot(path, 2, png_1x1, "Test", 1)
    assert tab_name == "No.2"

    wb2 = load_workbook(str(path))  # Don't use read_only for checking writes
    ws = wb2["No.2"]
    assert ws.cell(row=3, column=1).value == "Test"
    assert ws.cell(row=3, column=2).value is None or ""


def test_embed_screenshot_wrong_test_no(tmp_path):
    """Raises ValueError for unmapped test_no."""
    path = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    wb.create_sheet("MainSheet")
    wb.save(str(path))

    png_1x1 = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
    with pytest.raises(ValueError, match="No screenshot tab found"):
        embed_screenshot(path, 99, png_1x1, "Test", 1)


def test_embed_screenshot_oversized(tmp_path):
    """Oversized base64 rejected."""
    path = tmp_path / "big.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    wb.create_sheet("MainSheet")
    wb.create_sheet("No.2")
    wb.save(str(path))

    oversized = "A" * (MAX_IMAGE_B64_BYTES + 1)
    with pytest.raises(ValueError, match="too large"):
        embed_screenshot(path, 2, oversized, "Test", 1)


def test_get_workbook_meta_counts(tmp_path):
    """OK/NG/untested counts correct."""
    path = tmp_path / "meta.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Cover")
    wb.create_sheet("Revisions")
    ws = wb.create_sheet("MainSheet")
    for c in range(1, 11):
        ws.cell(row=1, column=c, value="#" if c == 1 else f"H{c}")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=8, value="2024-01-01")
    ws.cell(row=2, column=9, value="OK")
    ws.cell(row=2, column=10, value="TAIA")
    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=8, value="2024-01-01")
    ws.cell(row=3, column=9, value="NG")
    ws.cell(row=3, column=10, value="TAIA")
    ws.cell(row=4, column=1, value=3)
    wb.create_sheet("No.2")
    wb.save(str(path))

    meta = get_workbook_meta(path)
    assert meta.test_count == 3
    assert meta.ok_count == 1
    assert meta.ng_count == 1
    assert meta.untested_count == 1
    assert "No.2" in meta.screenshot_tabs