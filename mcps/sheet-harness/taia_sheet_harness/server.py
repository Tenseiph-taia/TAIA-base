import logging
import os
import re
from pathlib import Path
from typing import Optional

import base64
from filelock import FileLock, Timeout
from fastmcp import FastMCP
from openpyxl import load_workbook

from .models import (
    ImportWorkbookInput, WriteResultInput, EmbedScreenshotInput,
    WorkbookMeta, OkNg, TestCase,
    ReadSheetInput, ReadSheetOutput,
    WriteCellInput, WriteRangeInput,
    SheetListOutput, SheetInfo,
    CreateSheetInput, DeleteSheetInput,
    GetWorkbookSummaryOutput,
)
from .workbook import (
    import_workbook, read_test_cases, write_result,
    embed_screenshot, get_workbook_meta,
    read_sheet, write_cell, write_range,
    create_sheet, delete_sheet, get_sheet_info,
    get_main_sheet_info, get_workbook_summary,
    parse_tab_map, _find_main_sheet, HEADER_ROW, Col,
)

logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
)
log = logging.getLogger(__name__)


# -- Exception hierarchy ---------------------------------------------------------

class SheetHarnessError(Exception):
    """Base for all known, user-facing errors."""

class ValidationError(SheetHarnessError):
    """Bad input from the caller — equivalent to HTTP 400."""

class StorageError(SheetHarnessError):
    """Disk / OS error — equivalent to HTTP 503."""


def _user_error(msg: str) -> dict:
    return {"ok": False, "error": msg, "error_type": "validation"}


def _lock_error() -> dict:
    return {"ok": False, "error": "Could not acquire file lock — another operation is in progress. Retry in a moment.", "error_type": "lock_timeout"}


def _storage_error(exc: Exception) -> dict:
    log.error("Storage error: %s", exc, exc_info=True)
    return {"ok": False, "error": "Storage error. Check server logs.", "error_type": "server"}


# -- Configuration --------------------------------------------------------------

WORKBOOKS_DIR = Path(os.getenv("WORKBOOKS_DIR", "/workbooks"))
LIBRECHAT_UPLOADS_DIR = Path(os.getenv("LIBRECHAT_UPLOADS_DIR", "/librechat-uploads"))
LOCK_TIMEOUT  = int(os.getenv("LOCK_TIMEOUT_SECONDS", "10"))
MAX_WORKBOOKS = int(os.getenv("MAX_WORKBOOKS", "100"))
MAX_ROW       = 65_536

# MongoDB connection
from pymongo import MongoClient

MONGO_URI = os.getenv("MONGO_URI", "mongodb://mongodb:27017")
MONGO_DB = os.getenv("MONGO_DB", "LibreChat")   # matches your file records

_db_instance = None

def _get_db():
    global _db_instance
    if _db_instance is not None:
        return _db_instance
    try:
        client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=2000)
        client.admin.command('ping')
        _db_instance = client[MONGO_DB]
        return _db_instance
    except Exception:
        log.warning("MongoDB unavailable; file listing will be unavailable.")
        return None

WORKBOOKS_DIR.mkdir(parents=True, exist_ok=True)

mcp = FastMCP(
    name="taia-sheet-harness",
    instructions=(
        "TAIA Excel test harness for LLM orchestration. "
        "Import user Excel files, read test cases, write OK/NG results, embed screenshots. "
        "The main test sheet is always at index 2 (third tab). "
        "Screenshot tabs are pre-created by the user (e.g. 'No.2', 'No.8-10'). "
        "Never create screenshot tabs — they must already exist. "
        "Never touch columns K, L, M (human review). "
        "Always call get_test_cases first to obtain valid row numbers."
    ),
)


# -- Helpers -------------------------------------------------------------------

_SAFE_NAME = re.compile(r"^[a-zA-Z0-9_\-]+$")


def _resolve(workbook_id: str) -> Path:
    if not _SAFE_NAME.match(workbook_id):
        raise ValueError(f"Invalid workbook_id '{workbook_id}': use only letters, digits, _ and -")
    if len(workbook_id) > 128:
        raise ValueError("workbook_id exceeds 128 character limit.")
    resolved = (WORKBOOKS_DIR / f"{workbook_id}.xlsx").resolve()
    try:
        resolved.relative_to(WORKBOOKS_DIR.resolve())
    except ValueError:
        raise ValueError("workbook_id resolves outside the workbooks directory.")
    return resolved


def _lock_path(path: Path) -> Path:
    return path.with_suffix(".lock")


def _require_exists(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(
            f"Workbook '{path.stem}' not found. "
            f"Call import_workbook first or check list_workbooks."
        )


# -- MCP Tools ----------------------------------------------------------------

@mcp.tool(
    description=(
        "List the current user's uploaded .xlsx files. "
        "Use this when the user asks to 'import' a file but does not provide the filename. "
        "Accepts the current user's display name (as shown in the system prompt under 'Current User'). "
        "If exactly one file is found, call import_workbook_tool with that filename automatically. "
        "If multiple, present the list to the user and ask which to import."
    )
)
def list_my_xlsx_files(user_name: str) -> dict:
    """Look up the user by display name, then return their .xlsx file metadata."""
    if not user_name or not user_name.strip():
        return {"ok": False, "error": "User name is required."}

    db = _get_db()
    if db is None:
        return {"ok": False, "error": "Database temporarily unavailable. Please try again later."}

    # 1. Find the user OID
    user_doc = db["users"].find_one({"name": user_name.strip()})
    if not user_doc:
        return {"ok": False, "error": f"No user found with name '{user_name}'."}

    user_oid = user_doc["_id"]

    # 2. Query files for this user, .xlsx only
    try:
        cursor = db["files"].find({
            "user": user_oid,
            "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }).sort("createdAt", -1)
    except Exception as exc:
        log.error("list_my_xlsx_files: MongoDB query failed: %s", exc)
        return {"ok": False, "error": "Database query failed."}

    files = []
    for doc in cursor:
        files.append({
            "filename": doc.get("filename", "unknown.xlsx"),
            "file_id": doc.get("file_id", ""),
            "created_at": doc.get("createdAt").isoformat() if doc.get("createdAt") else None,
        })

    return {"ok": True, "count": len(files), "files": files}


@mcp.tool(
    description=(
        "Import a user's Excel test plan from an uploaded file. "
        "The user must first upload the .xlsx file in TAIA chat, then call this tool with the exact filename. "
        "The file is read directly from the LibreChat uploads directory. "
        "Returns workbook_id and test case count."
    )
)
def import_workbook_tool(
    filename: str,
    workbook_id: str = None,
    overwrite: bool = False,
) -> dict:
    if not filename.lower().endswith(".xlsx"):
        return _user_error("Only .xlsx files are supported.")

    # Try exact match first
    source = LIBRECHAT_UPLOADS_DIR / filename
    if not source.exists():
        # LibreChat prefixes files with UUID__ — find most recent match by suffix
        suffix = f"__{filename}"
        matches = sorted(
            [f for f in LIBRECHAT_UPLOADS_DIR.rglob("*") if f.name.endswith(suffix)],
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        )
        if not matches:
            return _user_error(
                f"File '{filename}' not found in uploads. "
                "Please check that the file was uploaded successfully and that you typed the exact filename (including .xlsx)."
            )
        source = matches[0]

    # Derive clean workbook_id from original filename — never expose the UUID prefix
    wid = workbook_id or re.sub(r"[^a-zA-Z0-9_\-]", "_", Path(filename).stem)[:128]

    try:
        path = _resolve(wid)
    except ValueError as exc:
        return _user_error(str(exc))

    if path.exists() and not overwrite:
        return _user_error(f"Workbook '{wid}' already exists. Set overwrite=true to replace it.")

    if len(list(WORKBOOKS_DIR.glob("*.xlsx"))) >= MAX_WORKBOOKS and not path.exists():
        return _user_error(f"Maximum workbook limit ({MAX_WORKBOOKS}) reached.")

    try:
        import shutil
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            shutil.copy2(source, path)
            file_bytes = path.read_bytes()
            meta = import_workbook(path, file_bytes)
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)
    except Exception as exc:
        log.error("Unexpected error during import: %s", exc, exc_info=True)
        return {"ok": False, "error": f"Import failed: {exc}", "error_type": "server"}

    log.info("import_workbook workbook_id=%s source=%s", wid, source)
    return {
        "ok": True,
        "workbook_id": meta.workbook_id,
        "filename": meta.filename,
        "main_sheet_name": meta.main_sheet_name,
        "test_count": meta.test_count,
        "ok_count": meta.ok_count,
        "ng_count": meta.ng_count,
        "untested_count": meta.untested_count,
        "screenshot_tabs": meta.screenshot_tabs,
        "message": f"Workbook imported with {meta.test_count} test cases.",
    }




@mcp.tool(
    description=(
        "Read all test cases from the imported workbook's main sheet (index 2). "
        "Each test case includes its row number — use this row number when writing results. "
        "Call this before any write operation. "
        "Rows without a numeric value in column A (test number) are skipped (section headers)."
    )
)
def get_test_cases(
    workbook_id: str,
    untested_only: bool = False,
) -> dict:
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))

    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            cases = read_test_cases(path)
    except Timeout:
        return _lock_error()
    except OSError as exc:
        return _storage_error(exc)

    if untested_only:
        cases = [c for c in cases if not c.test_okng]

    return {
        "ok": True,
        "workbook_id": workbook_id,
        "count": len(cases),
        "test_cases": [c.model_dump() for c in cases],
    }


@mcp.tool(
    description=(
        "Write OK or NG result to a specific test case row. "
        "Also writes today's date and 'TAIA' as the tester. "
        "Row number must come from get_test_cases. "
        "OK/NG values are exactly 'OK' or 'NG' (case-sensitive)."
    )
)
def write_result_tool(
    workbook_id: str,
    row: int,
    ok_ng: OkNg,
    notes: str = "",
) -> dict:
    if err := _validate_row(row):
        return _user_error(err)

    try:
        path = _resolve(workbook_id)
    except ValueError as exc:
        return _user_error(str(exc))

    try:
        _require_exists(path)
    except FileNotFoundError as exc:
        return _user_error(str(exc))

    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            write_result(path, row, ok_ng, notes or "")
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)

    log.info("write_result workbook=%s row=%d ok_ng=%s", workbook_id, row, ok_ng.value)
    return {
        "ok": True,
        "workbook_id": workbook_id,
        "row": row,
        "ok_ng": ok_ng.value,
        "message": f"Row {row} updated with {ok_ng.value}.",
    }


def _validate_row(row: int) -> Optional[str]:
    if row < 2:
        return "Row must be >= 2 (row 1 is the header)."
    if row > MAX_ROW:
        return f"Row {row} exceeds the maximum of {MAX_ROW}."
    return None


@mcp.tool(
    description=(
        "Embed a screenshot into the pre-existing screenshot tab for a test case. "
        "Tab must already exist in the workbook (user-created). "
        "If no tab exists for the test number, returns an error — do not create new tabs. "
        "Multiple screenshots for the same test stack vertically. "
        "image_b64 must be base64-encoded PNG or JPEG. Maximum size: 10 MB base64."
    )
)
def embed_screenshot_tool(
    workbook_id: str,
    test_no: int,
    image_b64: str,
    caption: str = "",
    step_number: int = 1,
) -> dict:
    if err := _validate_row(test_no):
        return _user_error(err)

    try:
        path = _resolve(workbook_id)
    except ValueError as exc:
        return _user_error(str(exc))

    try:
        _require_exists(path)
    except FileNotFoundError as exc:
        return _user_error(str(exc))

    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            tab_name = embed_screenshot(path, test_no, image_b64, caption or "", step_number)
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)

    log.info("embed_screenshot workbook=%s test_no=%d tab=%s", workbook_id, test_no, tab_name)
    return {
        "ok": True,
        "workbook_id": workbook_id,
        "test_no": test_no,
        "tab_name": tab_name,
        "message": f"Screenshot embedded in tab '{tab_name}'.",
    }


@mcp.tool(
    description=(
        "Return OK/NG/untested counts and screenshot tab info for a workbook. "
        "Computed from the live main sheet data."
    )
)
def get_workbook_summary_tool(
    workbook_id: str,
) -> dict:
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))

    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            meta = get_workbook_meta(path)
    except Timeout:
        return _lock_error()
    except OSError as exc:
        return _storage_error(exc)

    return {
        "ok": True,
        "workbook_id": meta.workbook_id,
        "filename": meta.filename,
        "main_sheet_name": meta.main_sheet_name,
        "test_count": meta.test_count,
        "ok_count": meta.ok_count,
        "ng_count": meta.ng_count,
        "untested_count": meta.untested_count,
        "screenshot_tabs": meta.screenshot_tabs,
    }


@mcp.tool(
    description="List all workbooks in the workbooks directory."
)
def list_workbooks_tool() -> dict:
    workbooks = []
    for p in sorted(WORKBOOKS_DIR.glob("*.xlsx")):
        stat = p.stat()
        workbooks.append({
            "workbook_id": p.stem,
            "size_bytes": stat.st_size,
            "modified_at": stat.st_mtime,
        })
    return {"ok": True, "count": len(workbooks), "workbooks": workbooks}


@mcp.tool(
    description=(
        "Read data from a sheet in the workbook. "
        "Returns list of dicts with column headers as keys, or list of lists. "
        "Supports reading specific ranges and skipping empty rows."
    )
)
def read_sheet_tool(
    workbook_id: str,
    sheet: str | int,
    start_row: int = 1,
    end_row: int | None = None,
    start_col: int = 1,
    end_col: int | None = None,
    skip_empty_rows: bool = True,
    return_dicts: bool = True,
) -> dict:
    """Read data from a sheet in the workbook."""
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))
    
    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            data, headers = read_sheet(
                path, sheet, start_row, end_row, start_col, end_col,
                skip_empty_rows, return_dicts
            )
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)
    
    return {
        "ok": True,
        "workbook_id": workbook_id,
        "data": data,
        "headers": headers,
        "count": len(data),
    }


@mcp.tool(
    description=(
        "Write a value to a specific cell in a sheet. "
        "Supports writing text, numbers, booleans, and formulas. "
        "Optional cell styling available."
    )
)
def write_cell_tool(
    workbook_id: str,
    sheet: str | int,
    row: int,
    column: int,
    value: str,
    is_formula: bool = False,
    background_color: str | None = None,
    font_color: str | None = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: int | None = None,
) -> dict:
    """Write a value to a specific cell."""
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))
    
    from .models import CellStyle
    
    style = None
    if any([background_color, font_color, bold, italic, underline, font_size]):
        style = CellStyle(
            background_color=background_color,
            font_color=font_color,
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,
        )
    
    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            write_cell(path, sheet, row, column, value, is_formula, style)
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)
    
    return {
        "ok": True,
        "workbook_id": workbook_id,
        "sheet": sheet,
        "row": row,
        "column": column,
        "message": f"Cell ({row}, {column}) written.",
    }


@mcp.tool(
    description=(
        "Write a 2D array of values to a range of cells. "
        "Supports header row with special styling. "
        "Useful for writing tables of data."
    )
)
def write_range_tool(
    workbook_id: str,
    sheet: str | int,
    start_row: int,
    start_col: int,
    data: list[list],
    has_header: bool = False,
    background_color: str | None = None,
    font_color: str | None = None,
) -> dict:
    """Write a 2D array of values to a range of cells."""
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))
    
    from .models import CellStyle
    
    style = None
    if any([background_color, font_color]):
        style = CellStyle(
            background_color=background_color,
            font_color=font_color,
        )
    
    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            write_range(path, sheet, start_row, start_col, data, has_header, style)
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)
    
    rows_written = len(data) if data else 0
    cols_written = len(data[0]) if data and data[0] else 0
    
    return {
        "ok": True,
        "workbook_id": workbook_id,
        "rows_written": rows_written,
        "cols_written": cols_written,
        "message": f"Wrote {rows_written} rows x {cols_written} columns.",
    }


@mcp.tool(
    description=(
        "Create a new sheet in the workbook. "
        "Optional position to insert at specific location."
    )
)
def create_sheet_tool(
    workbook_id: str,
    name: str,
    position: int | None = None,
) -> dict:
    """Create a new sheet in the workbook."""
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))
    
    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            sheet_name = create_sheet(path, name, position)
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)
    
    return {
        "ok": True,
        "workbook_id": workbook_id,
        "sheet_name": sheet_name,
        "message": f"Sheet '{sheet_name}' created.",
    }


@mcp.tool(
    description=(
        "Delete a sheet from the workbook. "
        "Cannot delete the main sheet or cover/revisions sheets."
    )
)
def delete_sheet_tool(
    workbook_id: str,
    sheet: str | int,
) -> dict:
    """Delete a sheet from the workbook."""
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))
    
    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            delete_sheet(path, sheet)
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)
    
    return {
        "ok": True,
        "workbook_id": workbook_id,
        "message": f"Sheet '{sheet}' deleted.",
    }


@mcp.tool(
    description=(
        "Get information about all sheets in the workbook. "
        "Returns sheet names, dimensions, and identifies the main sheet."
    )
)
def get_sheets_info_tool(
    workbook_id: str,
) -> dict:
    """Get information about all sheets in the workbook."""
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))
    
    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            sheets = get_sheet_info(path)
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)
    
    # Get main sheet name
    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            wb = load_workbook(str(path), read_only=True)
            main_sheet = _find_main_sheet(wb).title
            wb.close()
    except Exception:
        main_sheet = None
    
    return {
        "ok": True,
        "workbook_id": workbook_id,
        "sheet_count": len(sheets),
        "main_sheet_name": main_sheet,
        "sheets": [s.model_dump() for s in sheets],
    }


@mcp.tool(
    description=(
        "Get comprehensive summary of a workbook including metadata and file stats."
    )
)
def get_workbook_summary_extended_tool(
    workbook_id: str,
) -> dict:
    """Get comprehensive summary of a workbook."""
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))
    
    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            summary = get_workbook_summary(path)
    except Timeout:
        return _lock_error()
    except ValueError as exc:
        return _user_error(str(exc))
    except OSError as exc:
        return _storage_error(exc)
    
    return summary.model_dump()


@mcp.tool(
    description="Validate that a workbook has the correct structure and all required sheets."
)
def validate_workbook_tool(
    workbook_id: str,
) -> dict:
    try:
        path = _resolve(workbook_id)
        _require_exists(path)
    except (ValueError, FileNotFoundError) as exc:
        return _user_error(str(exc))

    try:
        with FileLock(str(_lock_path(path)), timeout=LOCK_TIMEOUT):
            from openpyxl import load_workbook as load_wb
            wb = load_wb(str(path), read_only=True)
            sheet_names = wb.sheetnames

            if len(sheet_names) < 3:
                return _user_error(f"Workbook has only {len(sheet_names)} sheet(s). Expected at least 3 (Cover, Revisions, main test sheet).")

            main_sheet_ws = _find_main_sheet(wb)
            main_sheet = main_sheet_ws.title

            # Header already validated by _find_main_sheet

            # Check for screenshot tabs
            screenshot_tabs = parse_tab_map(wb, main_sheet)

    except Timeout:
        return _lock_error()
    except OSError as exc:
        return _storage_error(exc)

    return {
        "ok": True,
        "workbook_id": workbook_id,
        "main_sheet": main_sheet,
        "screenshot_tabs": sorted(set(screenshot_tabs.values())),
        "message": "Workbook structure is valid.",
    }

@mcp.tool(
    description=(
        "Get the download URL for a completed workbook. "
        "Use this when the user asks to download their test results. "
        "Returns a direct HTTP link the user can open to download the .xlsx file."
    )
)
def get_download_url(workbook_id: str) -> dict:
    try:
        path = _resolve(workbook_id)
    except ValueError:
        return {"ok": False, "error": "Invalid workbook_id"}

    if not path.exists():
        return {"ok": False, "error": "Workbook not found"}

    # Use the server's hostname or IP from environment
    host = os.getenv("HOST", "0.0.0.0")
    download_port = int(os.getenv("DOWNLOAD_PORT", "8007"))
    # Prefer a public-facing hostname if set (e.g. TAIA_SERVER_HOST)
    public_host = os.getenv("TAIA_SERVER_HOST", host)
    url = f"http://{public_host}:{download_port}/download/{workbook_id}"
    return {"ok": True, "url": url, "filename": f"{workbook_id}.xlsx"}


# -- Entrypoint ------------------------------------------------------------------

def main() -> None:
    port = int(os.getenv("PORT_MCP", os.getenv("PORT", "8006")))
    mcp.run(transport="sse", host="0.0.0.0", port=port)