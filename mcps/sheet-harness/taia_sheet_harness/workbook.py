from __future__ import annotations

import base64
import io
import logging
import shutil
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Side, Border, Color
from openpyxl.styles.fills import PatternFill as PatternFillStyle

from .models import (
    Col, DATA_START, HEADER_ROW,
    OkNg, TestCase, WorkbookMeta, WriteResultInput,
    CellStyle, CellValue, SheetInfo, GetWorkbookSummaryOutput,
)

log = logging.getLogger(__name__)

# ── Limits ─────────────────────────────────────────────────────────────────────

MAX_TEXT_LEN = 32_767  # Excel single-cell character limit
MAX_IMAGE_B64_BYTES = 10_485_760  # 10 MB of base64 input

_BACKUP_RETAIN = 5  # Number of backups to keep per workbook

# ── Palette ────────────────────────────────────────────────────────────────────

_C = {
    "ok_bg": "C6EFCE",
    "ok_fg": "276221",
    "ng_bg": "FFC7CE",
    "ng_fg": "9C0006",
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
    "border": "BFBFBF",
}

_thin = Side(style="thin", color=_C["border"])
_center = Font(name="Arial", size=10)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(color: str = "000000", bold: bool = False, size: int = 10) -> Font:
    return Font(name="Arial", color=color, bold=bold, size=size)


def _backup(path: Path) -> None:
    """Create a timestamped backup of the workbook. Keeps last _BACKUP_RETAIN backups."""
    if not path.exists():
        return
    backup_dir = path.parent / ".backups"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    shutil.copy2(path, backup_dir / f"{path.stem}_{ts}.xlsx")
    existing = sorted(backup_dir.glob(f"{path.stem}_*.xlsx"))
    for old in existing[:-_BACKUP_RETAIN]:
        old.unlink(missing_ok=True)


def _atomic_save(wb, path: Path) -> None:
    """Write to a temp file then atomically rename — prevents corrupt workbooks on crash."""
    import shutil as sh
    if path.exists():
        _backup(path)
    tmp = path.with_suffix(".xlsx.tmp")
    try:
        wb.close()
        wb.save(str(tmp))
        sh.move(str(tmp), str(path))
    except Exception:
        tmp.unlink(missing_ok=True)
        raise


# ── Tab name parsing ───────────────────────────────────────────────────────────

def parse_tab_map(wb: Workbook, main_sheet_name: str) -> dict[int, str]:
    """
    Read all tab names from the main sheet index + 1 onwards.
    For "No.2" → {2: "No.2"}
    For "No.8-10" → {8: "No.8-10", 9: "No.8-10", 10: "No.8-10"}
    Return dict mapping test_no (int) to tab name (str).
    Handles both "No.2" and "No.8-10" formats.
    Ignores tabs that don't match the pattern.
    """
    result: dict[int, str] = {}
    main_idx = wb.sheetnames.index(main_sheet_name)
    for name in wb.sheetnames[main_idx + 1:]:
        name_strip = name.strip()
        if name_strip.startswith("No."):
            range_part = name_strip[3:]  # After "No."
            if "-" in range_part:
                try:
                    start, end = map(int, range_part.split("-"))
                    for n in range(start, end + 1):
                        result[n] = name_strip
                except ValueError:
                    continue
            else:
                try:
                    n = int(range_part)
                    result[n] = name_strip
                except ValueError:
                    continue
    return result


# ── Sheet Discovery (Enhanced) ────────────────────────────────────────────────

def _find_main_sheet(wb) -> object:
    """
    Find the main test sheet by scanning for header value "#" in A1.
    Accepts "#", "No.", "No".
    Returns the worksheet object.
    Falls back to index 2 if no match found.
    Raises ValueError if no main sheet found.
    """
    # First pass: scan by header
    for ws in wb.worksheets:
        val = ws.cell(row=1, column=1).value
        if val is not None and str(val).strip() in ("#", "No.", "No"):
            return ws
    # Fallback: index 2 if it exists
    if len(wb.worksheets) > 2:
        log.warning("No header match found — falling back to sheet index 2: %s", wb.worksheets[2].title)
        return wb.worksheets[2]
    raise ValueError("Could not identify main test sheet. Workbook must have at least 3 sheets.")


def _find_sheet(wb, sheet_name_or_index: str | int | None = None):
    """
    Find and return a sheet by name or index.
    - If sheet_name_or_index is None: return main sheet
    - If sheet_name_or_index is int: return by 0-based index
    - If sheet_name_or_index is str: return by name
    """
    if sheet_name_or_index is None:
        # Return main sheet for backward compatibility
        return _find_main_sheet(wb)
    
    if isinstance(sheet_name_or_index, int):
        # Index-based access
        if 0 <= sheet_name_or_index < len(wb.worksheets):
            return wb.worksheets[sheet_name_or_index]
        raise ValueError(f"Sheet index {sheet_name_or_index} out of range. Workbook has {len(wb.worksheets)} sheets.")
    
    # Name-based access
    if sheet_name_or_index in wb.sheetnames:
        return wb[sheet_name_or_index]
    raise ValueError(f"Sheet '{sheet_name_or_index}' not found. Available sheets: {wb.sheetnames}")


def get_sheet_index(wb, sheet_name_or_index: str | int | None = None) -> int:
    """
    Get the 0-based index of a sheet by name or index.
    """
    if sheet_name_or_index is None:
        # Main sheet
        main_ws = _find_main_sheet(wb)
        return wb.worksheets.index(main_ws)
    
    if isinstance(sheet_name_or_index, int):
        return sheet_name_or_index
    
    return wb.sheetnames.index(sheet_name_or_index)


# ── Import ─────────────────────────────────────────────────────────────────────

def import_workbook(path: Path, file_bytes: bytes) -> WorkbookMeta:
    """
    Write file_bytes to path atomically.
    Open with openpyxl, validate tab count >= 3.
    Parse tab map. Read main sheet. Return WorkbookMeta.
    """
    _atomic_save_bytes(path, file_bytes)
    return get_workbook_meta(path)


def _atomic_save_bytes(path: Path, file_bytes: bytes) -> None:
    """Write bytes to a temp file then atomically rename."""
    if path.exists():
        _backup(path)
    tmp = path.with_suffix(".xlsx.tmp")
    try:
        tmp.write_bytes(file_bytes)
        tmp.replace(path)
    except Exception:
        tmp.unlink(missing_ok=True)
        raise


# ── Reading ────────────────────────────────────────────────────────────────────

def read_test_cases(path: Path) -> list[TestCase]:
    """
    Open workbook (data_only=True).
    Get ws = _find_main_sheet(wb).
    Iterate from DATA_START to max_row.
    Skip any row where Col A is None or not castable to int — these are section header rows.
    Only rows with a numeric value in Col A are test cases.
    Resolve screenshot_tab from tab_map for each test_no.
    Return list[TestCase].
    """
    wb = load_workbook(str(path), data_only=True)
    if len(wb.worksheets) < 3:
        return []

    main_sheet = _find_main_sheet(wb)
    tab_map = parse_tab_map(wb, main_sheet.title)
    cases: list[TestCase] = []

    for row_idx in range(DATA_START, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=Col.NO).value

        # Skip if Col A is empty or not an integer (section headers)
        if cell_val is None:
            continue
        try:
            test_no = int(cell_val)
        except (ValueError, TypeError):
            continue

        prerequisite = _strval(ws.cell(row=row_idx, column=Col.PREREQ).value)
        test_detail = _strval(ws.cell(row=row_idx, column=Col.TEST_DETAIL).value)
        expected_result = _strval(ws.cell(row=row_idx, column=Col.EXPECTED).value)
        data_no = _strval(ws.cell(row=row_idx, column=Col.DATA_NO).value)
        remarks = _strval(ws.cell(row=row_idx, column=Col.REMARKS).value)
        num_cases = _strval(ws.cell(row=row_idx, column=Col.NUM_CASES).value)

        test_date = _strval(ws.cell(row=row_idx, column=Col.TEST_DATE).value)
        test_okng = _strval(ws.cell(row=row_idx, column=Col.TEST_OKNG).value)

        screenshot_tab = tab_map.get(test_no)

        cases.append(
            TestCase(
                row=row_idx,
                test_no=test_no,
                prerequisite=prerequisite,
                test_detail=test_detail,
                expected_result=expected_result,
                data_no=data_no,
                remarks=remarks,
                num_cases=num_cases,
                test_date=test_date,
                test_okng=test_okng,
                screenshot_tab=screenshot_tab,
            )
        )

    return cases


def _strval(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def read_sheet(path: Path, sheet_name_or_index: str | int | None = None,
               start_row: int = 1, end_row: int | None = None,
               start_col: int = 1, end_col: int | None = None,
               skip_empty_rows: bool = True,
               return_dicts: bool = True) -> tuple[list[dict] | list[list], list[str] | None]:
    """
    Read data from a specific sheet.
    Returns (data, headers) where data is list of dicts (if return_dicts=True) or list of lists.
    """
    wb = load_workbook(str(path), data_only=True)
    ws = _find_sheet(wb, sheet_name_or_index)
    
    # Determine sheet index
    sheet_index = wb.worksheets.index(ws)
    sheet_name = ws.title
    
    # Adjust row/col indices for 0-based internal, 1-based API
    actual_start_row = max(1, start_row)
    actual_end_row = end_row if end_row else ws.max_row
    actual_start_col = max(1, start_col)
    actual_end_col = end_col if end_col else ws.max_column
    
    rows = range(actual_start_row, min(actual_end_row + 1, ws.max_row + 1))
    cols = range(actual_start_col, min(actual_end_col + 1, ws.max_column + 1))
    
    # Read data
    data = []
    headers = None
    
    for row_idx in rows:
        row_data = []
        for col_idx in cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(_cell_value(cell))
        
        # Skip empty rows if requested
        if skip_empty_rows and all(v is None or v == "" for v in row_data):
            continue
        
        if row_idx == actual_start_row and return_dicts:
            # First row as headers
            headers = row_data
        else:
            if return_dicts and headers:
                # Convert to dict
                row_dict = {}
                for i, header in enumerate(headers):
                    col_num = actual_start_col + i
                    row_dict[header if header else f"Col{col_num}"] = row_data[i]
                data.append(row_dict)
            else:
                data.append(row_data)
    
    wb.close()
    return data, headers


def _cell_value(cell: Cell) -> Any:
    """Extract the value from a cell, handling different types."""
    if cell.value is None:
        return None
    if isinstance(cell.value, str):
        return cell.value.strip()
    return cell.value


# ── Writing ────────────────────────────────────────────────────────────────────

def write_result(path: Path, row: int, ok_ng: OkNg, notes: str = "") -> None:
    """
    Load workbook (not data_only).
    Write to ws = _find_main_sheet(wb):
      - Col H (TEST_DATE): datetime.now().strftime("%Y-%m-%d")
      - Col I (TEST_OKNG): ok_ng.value — "OK" in green fill, "NG" in red fill
      - Col J (TEST_TESTER): "TAIA"
      - If notes provided: append to Col F (REMARKS)
    Atomic save.
    """
    wb = load_workbook(str(path))
    ws = _find_main_sheet(wb)

    today = datetime.now().strftime("%Y-%m-%d")
    ws.cell(row=row, column=Col.TEST_DATE, value=today)
    ws.cell(row=row, column=Col.TEST_OKNG, value=ok_ng.value)
    ws.cell(row=row, column=Col.TEST_TESTER, value="TAIA")

    # Styling for OK/NG
    if ok_ng == OkNg.OK:
        ws.cell(row=row, column=Col.TEST_OKNG).fill = _fill(_C["ok_bg"])
        ws.cell(row=row, column=Col.TEST_OKNG).font = _font(_C["ok_fg"], bold=True)
    else:
        ws.cell(row=row, column=Col.TEST_OKNG).fill = _fill(_C["ng_bg"])
        ws.cell(row=row, column=Col.TEST_OKNG).font = _font(_C["ng_fg"], bold=True)

    ws.cell(row=row, column=Col.TEST_DATE).font = _font(size=9)
    ws.cell(row=row, column=Col.TEST_TESTER).font = _font(bold=True, size=10)

    if notes:
        current = _strval(ws.cell(row=row, column=Col.REMARKS).value)
        new_val = f"{current}; {notes}" if current else notes
        ws.cell(row=row, column=Col.REMARKS, value=new_val[:MAX_TEXT_LEN])

    _atomic_save(wb, path)


def write_cell(path: Path, sheet_name_or_index: str | int | None,
               row: int, column: int, value: Any,
               is_formula: bool = False, style: CellStyle | None = None) -> None:
    """
    Write a single cell value with optional styling.
    """
    wb = load_workbook(str(path))
    ws = _find_sheet(wb, sheet_name_or_index)
    
    cell = ws.cell(row=row, column=column, value=value)
    
    # Handle formula
    if is_formula:
        cell.value = f"={value}"
    
    # Apply styling
    if style:
        _apply_cell_style(cell, style)
    
    _atomic_save(wb, path)


def write_range(path: Path, sheet_name_or_index: str | int | None,
                start_row: int, start_col: int,
                data: list[list], has_header: bool = False,
                style: CellStyle | None = None) -> None:
    """
    Write a 2D array of values to a range of cells.
    """
    wb = load_workbook(str(path))
    ws = _find_sheet(wb, sheet_name_or_index)
    
    # Write data
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, cell_value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Apply header styling if this is the header row
            if has_header and row_idx == start_row:
                cell.font = _font(color=_C["header_fg"], bold=True)
                cell.fill = _fill(_C["header_bg"])
            elif style:
                _apply_cell_style(cell, style)
    
    _atomic_save(wb, path)


def _apply_cell_style(cell: Cell, style: CellStyle) -> None:
    """Apply styling to a cell."""
    # Background color
    if style.background_color:
        cell.fill = _fill(style.background_color)
    
    # Font styling
    font_color = style.font_color or "000000"
    cell.font = _font(
        color=font_color,
        bold=style.bold,
        size=style.font_size or 10
    )
    cell.font = Font(
        name="Arial",
        color=font_color,
        bold=style.bold,
        italic=style.italic,
        underline="single" if style.underline else None,
        size=style.font_size or 10
    )


# ── Screenshots ────────────────────────────────────────────────────────────────

def embed_screenshot(path: Path, test_no: int, image_b64: str, caption: str, step_number: int) -> str:
    """
    Load workbook.
    Get tab_name from tab_map for test_no — raise ValueError if not found.
    Get ws = wb[tab_name].
    Find next available row to place screenshot (scan down from row 3 for empty anchor point).
    Write caption text to that row, Col A.
    Embed openpyxl Image anchored at Col B of the next row.
    Scale image to max width 500px maintaining aspect ratio.
    Set row height to fit.
    Atomic save.
    Return tab_name as confirmation.
    """
    if len(image_b64) > MAX_IMAGE_B64_BYTES:
        raise ValueError(
            f"Image data too large ({len(image_b64):,} bytes base64). "
            f"Maximum is {MAX_IMAGE_B64_BYTES:,} bytes."
        )

    wb = load_workbook(str(path))
    main_sheet = _find_main_sheet(wb)
    tab_map = parse_tab_map(wb, main_sheet.title)

    if test_no not in tab_map:
        raise ValueError(f"No screenshot tab found for test No. {test_no}")

    tab_name = tab_map[test_no]
    ws = wb[tab_name]

    # Find next available row (start from row 3, find first completely empty row in cols A-C)
    next_row = 3
    while True:
        has_content = False
        for col in range(1, 4):
            if ws.cell(row=next_row, column=col).value is not None:
                has_content = True
                break
        if not has_content:
            break
        next_row += 1

    # Write caption
    ws.cell(row=next_row, column=1, value=caption or f"Step {step_number}")
    ws.cell(row=next_row, column=1).font = _font(bold=True, size=10)

    # Decode and embed image
    try:
        img_data = base64.b64decode(image_b64, validate=True)
    except Exception as exc:
        raise ValueError(f"Invalid base64 image data: {exc}") from exc

    img_stream = io.BytesIO(img_data)
    xl_img = XLImage(img_stream)

    # Scale to max 500px width
    max_w = 500
    if xl_img.width and xl_img.width > max_w:
        scale = max_w / xl_img.width
        xl_img.width = int(xl_img.width * scale)
        xl_img.height = int((xl_img.height or 100) * scale)

    # Anchor at Col B, next row
    col_letter = "B"
    ws.add_image(xl_img, f"{col_letter}{next_row + 1}")

    # Set row height to fit image
    if xl_img.height:
        ws.row_dimensions[next_row + 1].height = xl_img.height * 0.75

    _atomic_save(wb, path)
    return tab_name


# ── Sheet Operations ───────────────────────────────────────────────────────────

def create_sheet(path: Path, name: str, position: int | None = None) -> str:
    """
    Create a new sheet in the workbook.
    Position is 0-based. If None, appends to end.
    """
    wb = load_workbook(str(path))
    
    # Create sheet
    if position is None:
        ws = wb.create_sheet(name)
    else:
        ws = wb.create_sheet(name, position)
    
    _atomic_save(wb, path)
    return ws.title


def delete_sheet(path: Path, sheet_name_or_index: str | int) -> None:
    """
    Delete a sheet from the workbook.
    """
    wb = load_workbook(str(path))
    
    # Delete sheet
    if isinstance(sheet_name_or_index, int):
        if 0 <= sheet_name_or_index < len(wb.worksheets):
            sheet_name = wb.worksheets[sheet_name_or_index].title
            del wb[sheet_name]
    else:
        del wb[sheet_name_or_index]
    
    _atomic_save(wb, path)


def get_sheet_info(path: Path) -> list[SheetInfo]:
    """
    Get information about all sheets in the workbook.
    """
    wb = load_workbook(str(path), read_only=True)
    
    info = []
    for i, ws in enumerate(wb.worksheets):
        info.append(SheetInfo(
            name=ws.title,
            index=i,
            rows=ws.max_row,
            columns=ws.max_column,
            is_main_sheet=(i == 0 or i == 2)  # Simplified - check header later
        ))
    
    wb.close()
    return info


def get_main_sheet_info(path: Path) -> SheetInfo | None:
    """
    Get information about the main sheet.
    """
    try:
        wb = load_workbook(str(path), read_only=True)
        ws = _find_main_sheet(wb)
        
        info = SheetInfo(
            name=ws.title,
            index=wb.worksheets.index(ws),
            rows=ws.max_row,
            columns=ws.max_column,
            is_main_sheet=True
        )
        
        wb.close()
        return info
    except Exception:
        return None


# ── Summary ────────────────────────────────────────────────────────────────────

def get_workbook_meta(path: Path) -> WorkbookMeta:
    """
    Read test cases, count OK/NG/untested.
    Return WorkbookMeta.
    """
    cases = read_test_cases(path)
    ok_count = sum(1 for c in cases if c.test_okng == "OK")
    ng_count = sum(1 for c in cases if c.test_okng == "NG")
    untested = sum(1 for c in cases if not c.test_okng)

    wb = load_workbook(str(path), read_only=True)
    main_sheet = _find_main_sheet(wb)
    main_sheet_name = main_sheet.title
    tab_map = parse_tab_map(wb, main_sheet_name)

    return WorkbookMeta(
        workbook_id=path.stem,
        filename=path.name,
        main_sheet_name=main_sheet_name,
        test_count=len(cases),
        ok_count=ok_count,
        ng_count=ng_count,
        untested_count=untested,
        screenshot_tabs=sorted(set(tab_map.values())),
    )


def get_workbook_summary(path: Path) -> GetWorkbookSummaryOutput:
    """
    Get comprehensive summary of workbook including metadata and file stats.
    """
    cases = []
    ok_count = 0
    ng_count = 0
    untested = 0
    main_sheet_name = None
    screenshot_tabs = []
    
    try:
        cases = read_test_cases(path)
        ok_count = sum(1 for c in cases if c.test_okng == "OK")
        ng_count = sum(1 for c in cases if c.test_okng == "NG")
        untested = sum(1 for c in cases if not c.test_okng)
    except Exception as e:
        log.warning(f"Could not read test cases: {e}")
    
    try:
        wb = load_workbook(str(path), read_only=True)
        main_sheet = _find_main_sheet(wb)
        main_sheet_name = main_sheet.title
        tab_map = parse_tab_map(wb, main_sheet_name)
        screenshot_tabs = sorted(set(tab_map.values()))
        sheet_count = len(wb.worksheets)
        wb.close()
    except Exception as e:
        log.warning(f"Could not read workbook metadata: {e}")
        wb.close()
        sheet_count = 0
    
    # Get file stats
    file_stat = os.stat(path) if path.exists() else None
    
    return GetWorkbookSummaryOutput(
        ok=True,
        workbook_id=path.stem,
        filename=path.name if path.exists() else "",
        sheet_count=sheet_count,
        main_sheet_name=main_sheet_name,
        test_count=len(cases),
        ok_count=ok_count,
        ng_count=ng_count,
        untested_count=untested,
        screenshot_tabs=screenshot_tabs,
        created_at=datetime.fromtimestamp(file_stat.st_ctime).isoformat() if file_stat else datetime.now().isoformat(),
        modified_at=datetime.fromtimestamp(file_stat.st_mtime).isoformat() if file_stat else datetime.now().isoformat(),
    )